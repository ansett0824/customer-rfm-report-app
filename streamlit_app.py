# 匯入 pandas：用來讀取與整理表格資料（例如 Excel）
import pandas as pd

# 匯入 numpy：常用於數值運算，這裡主要拿來處理圖表座標
import numpy as np

# 關閉 joblib 偵測 CPU 核心數的警告
import os
os.environ["LOKY_MAX_CPU_COUNT"] = "4"

# 匯入 matplotlib：用來畫圖表
import matplotlib.pyplot as plt

# 匯入 KMeans：用來做 K-means 分群分析
from sklearn.cluster import KMeans

# 匯入 Workbook：用來建立 Excel 活頁簿
from openpyxl import Workbook

# 匯入 Excel 樣式功能：字型、底色、對齊、框線
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 匯入 dataframe_to_rows：可以把 pandas 的表格資料一列一列寫進 Excel
from openpyxl.utils.dataframe import dataframe_to_rows

# 匯入 get_column_letter：用來把欄位數字轉成 Excel 欄位字母
from openpyxl.utils import get_column_letter

# 匯入圖片功能：把圖表圖片插入 Excel
from openpyxl.drawing.image import Image as XLImage

# 匯入字型管理：讓 matplotlib 可以顯示中文字
import matplotlib.font_manager as fm

# 匯入 Path：用比較方便的方式處理資料夾與檔案路徑
from pathlib import Path

# 匯入 datetime：用來取得目前時間，產生帶時間戳記的輸出檔名
from datetime import datetime
import tempfile
import streamlit as st




def generate_customer_report(
    sales_file,
    customer_info_file=None,
    analysis_months=None,
    vip_gap_ratio_threshold=0.5,
    vip_m_log_base=2.5
):
    """
    產生客群分析 Excel 報表，回傳輸出檔案路徑。
    """
    if analysis_months is None or len(analysis_months) == 0:
        analysis_months = [1, 2, 3]

    # 確保月份由小到大排序，避免 R/F 判定順序混亂
    analysis_months = sorted([int(m) for m in analysis_months])

    # =====================================
    # 0. 路徑設定
    # =====================================

    # 設定暫存資料夾，用於存放圖表圖片與輸出 Excel
    base_dir = Path(tempfile.mkdtemp(prefix="rfm_customer_report_"))

    # Streamlit 上傳檔案可直接給 pandas.read_excel 讀取
    source_file = sales_file
    if hasattr(source_file, "seek"):
        source_file.seek(0)

    # 客戶聯絡資料可以選填；若未上傳，聯絡欄位會保留空白
    customer_info_file = customer_info_file

    # 取得目前時間，產生輸出檔名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = base_dir / f"客群分析結果報表_{timestamp}.xlsx"

    # 暫存圖表圖片檔案路徑
    count_img = base_dir / "_tmp_count.png"
    rfm_img = base_dir / "_tmp_rfm.png"
    money_img = base_dir / "_tmp_money.png"
    contrib_img = base_dir / "_tmp_contrib.png"
    gap_img = base_dir / "_tmp_gap_method.png"

    # RFM 分析月份設定
    ANALYSIS_MONTHS = analysis_months

    # R、F 分數固定為 1~5 分；M 分數一般客戶為 1~5 分，VIP 客戶可超過 5 分
    RFM_SCORE_MIN = 1
    RFM_SCORE_MAX = 5


    # =====================================
    # 1. 設定中文字型
    # =====================================

    def setup_chinese_font():
        """
        設定 Matplotlib 中文字型。
        適用本機 Windows 與 Streamlit Cloud Linux 環境。

        注意：
        1. Windows 本機通常可使用 Microsoft JhengHei。
        2. Streamlit Community Cloud 建議另外新增 packages.txt，
           內容放：
           fonts-noto-cjk
           fonts-noto-cjk-extra
        """

        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm

        # 重新整理字型快取，避免雲端剛安裝字型後讀不到
        try:
            fm._load_fontmanager(try_read_cache=False)
        except Exception:
            pass

        candidate_fonts = [
            "Microsoft JhengHei",
            "Microsoft YaHei",
            "PMingLiU",
            "DFKai-SB",
            "Noto Sans CJK TC",
            "Noto Sans CJK JP",
            "Noto Sans CJK SC",
            "Noto Sans TC",
            "Noto Sans CJK",
            "SimHei",
            "Arial Unicode MS",
            "DejaVu Sans"
        ]

        available_fonts = {f.name for f in fm.fontManager.ttflist}

        for font_name in candidate_fonts:
            if font_name in available_fonts:
                plt.rcParams["font.family"] = font_name
                plt.rcParams["font.sans-serif"] = [font_name]
                plt.rcParams["axes.unicode_minus"] = False
                print(f"已使用中文字型：{font_name}")
                return font_name

        # 如果字型名稱抓不到，改用字型檔路徑搜尋 Noto CJK
        for font in fm.fontManager.ttflist:
            font_path = str(font.fname)
            font_name = str(font.name)

            if "NotoSansCJK" in font_path or "Noto Sans CJK" in font_name:
                try:
                    fm.fontManager.addfont(font_path)
                except Exception:
                    pass

                plt.rcParams["font.family"] = font_name
                plt.rcParams["font.sans-serif"] = [font_name]
                plt.rcParams["axes.unicode_minus"] = False
                print(f"已使用中文字型檔：{font_name}，路徑：{font_path}")
                return font_name

        plt.rcParams["axes.unicode_minus"] = False
        print("找不到中文字型，將使用系統預設字型，中文可能無法正確顯示。")
        return None


    setup_chinese_font()


    # =====================================
    # 2. 讀取與整理資料
    # =====================================

    # 讀取來源 Excel 檔案成為 DataFrame
    df = pd.read_excel(source_file)

    # 去除欄位名稱前後空白
    df.columns = df.columns.astype(str).str.strip()

    # 重新命名客戶欄位
    df = df.rename(columns={
        "客戶": "customer_id",
        "客戶簡稱": "customer_name"
    })

    # 建立月份欄位對應，例如：1月 -> M01、2月 -> M02
    month_rename_map = {}

    for m in ANALYSIS_MONTHS:
        original_col = f"{m}月"
        new_col = f"M{m:02d}"

        if original_col in df.columns:
            month_rename_map[original_col] = new_col

    df = df.rename(columns=month_rename_map)

    # 實際存在於 Excel 的月份欄位
    month_cols = [f"M{m:02d}" for m in ANALYSIS_MONTHS if f"M{m:02d}" in df.columns]

    if len(month_cols) == 0:
        raise ValueError("找不到可分析的月份欄位，請確認 Excel 是否有 1月、2月、3月等欄位。")

    # 保留客戶欄位與分析月份欄位
    df = df[["customer_id", "customer_name"] + month_cols].copy()

    # 將月份欄位轉成數字
    for col in month_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # 分析月份數
    analysis_month_count = len(month_cols)

    # 月份中文顯示名稱
    month_display_cols = [f"{int(col[1:])}月" for col in month_cols]


    # =====================================
    # 2-1. 讀取客戶聯絡資料並合併
    # =====================================

    def clean_text_key(value):
        """清理合併用文字鍵值。"""
        if pd.isna(value):
            return ""

        text = str(value).strip()

        if text.endswith(".0"):
            text = text[:-2]

        return text


    def clean_contact_value(value):
        """清理聯絡資料，避免電話被顯示為 223500.0。"""
        if pd.isna(value):
            return ""

        if isinstance(value, (int, np.integer)):
            return str(value)

        if isinstance(value, (float, np.floating)):
            if np.isnan(value):
                return ""

            if float(value).is_integer():
                return str(int(value))

            return str(value).strip()

        text = str(value).strip()

        if text.lower() == "nan":
            return ""

        if text.endswith(".0"):
            text = text[:-2]

        return text


    def merge_customer_contacts(main_df, contact_file):
        """
        從「客戶資料統計.xlsx」讀取聯絡人、聯絡電話、聯絡人手機。
        以主表客戶名稱對應客戶資料中的「客戶簡稱」；
        若找不到，再嘗試用「客戶全名」。
        沒有對應到的聯絡資訊保留空白。
        """
        result = main_df.copy()

        for col in ["聯絡人", "聯絡電話", "聯絡人手機"]:
            result[col] = ""

        if contact_file is None:
            print("未上傳客戶聯絡資料檔，聯絡欄位將保留空白。")
            return result

        try:
            if hasattr(contact_file, "seek"):
                contact_file.seek(0)
            contact_df = pd.read_excel(contact_file, sheet_name="客戶彙整")
        except Exception:
            if hasattr(contact_file, "seek"):
                contact_file.seek(0)
            contact_df = pd.read_excel(contact_file)

        contact_df.columns = contact_df.columns.astype(str).str.strip()

        for col in ["客戶簡稱", "客戶全名", "聯絡人", "聯絡電話", "聯絡人手機"]:
            if col not in contact_df.columns:
                contact_df[col] = ""

        contact_df = contact_df[[
            "客戶簡稱", "客戶全名",
            "聯絡人", "聯絡電話", "聯絡人手機"
        ]].copy()

        for col in ["客戶簡稱", "客戶全名"]:
            contact_df[col] = contact_df[col].apply(clean_text_key)

        for col in ["聯絡人", "聯絡電話", "聯絡人手機"]:
            contact_df[col] = contact_df[col].apply(clean_contact_value)

        result["_merge_customer_name"] = result["customer_name"].apply(clean_text_key)

        # 第一層：用客戶簡稱對應
        contact_short = contact_df[[
            "客戶簡稱", "聯絡人", "聯絡電話", "聯絡人手機"
        ]].copy()

        contact_short = contact_short[contact_short["客戶簡稱"] != ""]
        contact_short = contact_short.drop_duplicates(subset=["客戶簡稱"], keep="first")
        contact_short = contact_short.rename(columns={"客戶簡稱": "_merge_customer_name"})

        result = result.merge(
            contact_short,
            on="_merge_customer_name",
            how="left",
            suffixes=("", "_from_short")
        )

        for col in ["聯絡人", "聯絡電話", "聯絡人手機"]:
            result[col] = result[f"{col}_from_short"].fillna("").astype(str).str.strip()
            result = result.drop(columns=[f"{col}_from_short"])

        # 第二層：用客戶全名補
        contact_full = contact_df[[
            "客戶全名", "聯絡人", "聯絡電話", "聯絡人手機"
        ]].copy()

        contact_full = contact_full[contact_full["客戶全名"] != ""]
        contact_full = contact_full.drop_duplicates(subset=["客戶全名"], keep="first")
        contact_full = contact_full.rename(columns={"客戶全名": "_merge_customer_name"})

        result = result.merge(
            contact_full,
            on="_merge_customer_name",
            how="left",
            suffixes=("", "_from_full")
        )

        for col in ["聯絡人", "聯絡電話", "聯絡人手機"]:
            result[col] = result[col].fillna("").astype(str).str.strip().replace("nan", "")
            result[f"{col}_from_full"] = result[f"{col}_from_full"].fillna("").astype(str).str.strip().replace("nan", "")

            result[col] = np.where(
                result[col] == "",
                result[f"{col}_from_full"],
                result[col]
            )

            result = result.drop(columns=[f"{col}_from_full"])

        result = result.drop(columns=["_merge_customer_name"])

        matched_count = (
            result[["聯絡人", "聯絡電話", "聯絡人手機"]]
            .fillna("")
            .ne("")
            .any(axis=1)
            .sum()
        )

        print(f"已完成客戶聯絡資料合併，成功帶入聯絡資訊客戶數：{matched_count}")

        return result


    # 執行客戶聯絡資料合併
    df = merge_customer_contacts(df, customer_info_file)


    # =====================================
    # 3. 建立 RFM + 銷售貢獻度
    # =====================================

    # 負值修正為 0，避免退貨或折讓干擾正向 RFM 計算
    positive_sales = df[month_cols].clip(lower=0)

    # Monetary_raw：分析期間內原始合計銷售金額
    df["Monetary_raw"] = positive_sales.sum(axis=1)

    # 全部客戶總銷售金額
    total_sales = df["Monetary_raw"].sum()

    # 銷售貢獻度
    if total_sales > 0:
        df["銷售貢獻度"] = df["Monetary_raw"] / total_sales
    else:
        df["銷售貢獻度"] = 0

    # Frequency_raw：分析期間內有幾個月份有交易
    df["Frequency_raw"] = (positive_sales > 0).sum(axis=1)


    def calc_recency_raw_dynamic(row, month_columns):
        """
        Recency_raw：
        距離最近一次購買月份的間隔。
        0 代表最後分析月份有買，數字越大代表越久沒買。
        若完全無交易，回傳分析月份數。
        """
        for distance, col in enumerate(reversed(month_columns)):
            if row[col] > 0:
                return distance

        return len(month_columns)


    df["Recency_raw"] = df.apply(
        lambda row: calc_recency_raw_dynamic(row, month_cols),
        axis=1
    )


    def calc_r_score_dynamic(recency_raw, total_months):
        """
        R_score：1~5分
        最近有交易分數越高。
        """
        if total_months <= 1:
            return 5 if recency_raw == 0 else 1

        if recency_raw >= total_months:
            return 1

        score = 5 - np.floor((recency_raw / (total_months - 1)) * 4)
        score = int(score)

        return max(1, min(5, score))


    def calc_f_score_dynamic(frequency_raw, total_months):
        """
        F_score：1~5分
        有交易月份占比越高分數越高。
        """
        if total_months <= 0:
            return 1

        if frequency_raw <= 0:
            return 1

        score = int(np.ceil((frequency_raw / total_months) * 5))

        return max(1, min(5, score))


    df["R_score"] = df["Recency_raw"].apply(
        lambda x: calc_r_score_dynamic(x, analysis_month_count)
    )

    df["F_score"] = df["Frequency_raw"].apply(
        lambda x: calc_f_score_dynamic(x, analysis_month_count)
    )


    # =====================================
    # 3-1. Monetary 最大相鄰差額斷層法
    # =====================================

    def apply_monetary_gap_method(data, money_col="Monetary_raw"):
        """
        Monetary 改良式最大相鄰差額斷層法：
        1. 將所有客戶 Monetary 金額由小到大排序。
        2. 計算相鄰差額 = 目前金額 - 前一筆金額。
        3. 計算相鄰差距比例 = 相鄰差額 ÷ 前一筆金額。
        4. 只保留相鄰差距比例大於等於 vip_gap_ratio_threshold 的有效斷層。
        5. 若有多個有效斷層，選擇相鄰差額最大的斷層作為 VIP 判定位置。
        6. 採「斷層前一筆納入法」，將最大有效斷層的前一筆金額作為 VIP 門檻。
        """

        result = data.copy()
        valid_money = result.loc[result[money_col] > 0, money_col].copy()

        if len(valid_money) < 2:
            result["斷層起始門檻"] = np.nan
            result["斷層法分類"] = "一般客戶"
            result["金額排序"] = np.nan
            result["相鄰前一筆金額"] = np.nan
            result["相鄰差額"] = np.nan
            result["是否斷層起始點"] = "否"
            result["是否斷層極端客戶"] = "否"

            gap_info = {
                "gap_found": False,
                "gap_threshold": np.nan,
                "max_gap_amount": np.nan,
                "before_gap_amount": np.nan,
                "after_gap_amount": np.nan,
                "method_note": "有效 Monetary 金額不足 2 筆，未進行斷層判定。"
            }

            gap_table = result[["customer_id", "customer_name", money_col]].copy()
            return result, gap_info, gap_table

        gap_table = result[["customer_id", "customer_name", money_col]].copy()
        gap_table = gap_table.sort_values(money_col, ascending=True).reset_index(drop=False)

        gap_table["金額排序"] = np.arange(1, len(gap_table) + 1)
        gap_table["相鄰前一筆金額"] = gap_table[money_col].shift(1)
        gap_table["相鄰差額"] = gap_table[money_col] - gap_table["相鄰前一筆金額"]
        gap_table["相鄰差距比例"] = gap_table["相鄰差額"] / gap_table["相鄰前一筆金額"]

        valid_gap_table = gap_table[
            (gap_table[money_col] > 0) &
            (gap_table["相鄰前一筆金額"] > 0) &
            (gap_table["相鄰差距比例"] >= float(vip_gap_ratio_threshold))
        ].copy()

        if valid_gap_table.empty:
            result["斷層起始門檻"] = np.nan
            result["斷層法分類"] = "一般客戶"
            result["金額排序"] = np.nan
            result["相鄰前一筆金額"] = np.nan
            result["相鄰差額"] = np.nan
            result["是否斷層起始點"] = "否"
            result["是否斷層極端客戶"] = "否"

            gap_info = {
                "gap_found": False,
                "gap_threshold": np.nan,
                "max_gap_amount": np.nan,
                "before_gap_amount": np.nan,
                "after_gap_amount": np.nan,
                "method_note": f"沒有找到相鄰差距比例達 {float(vip_gap_ratio_threshold):.0%} 以上的有效斷層，因此未進行 VIP 門檻判定。"
            }

            return result, gap_info, gap_table

        selected = valid_gap_table.loc[valid_gap_table["相鄰差額"].idxmax()]

        before_gap_amount = selected["相鄰前一筆金額"]
        after_gap_amount = selected[money_col]
        gap_threshold = before_gap_amount
        max_gap_amount = selected["相鄰差額"]

        gap_ratio = max_gap_amount / before_gap_amount if before_gap_amount > 0 else np.nan

        method_note = (
            f"已偵測到符合 {float(vip_gap_ratio_threshold):.0%} 條件之最大金額斷層："
            f"由 {before_gap_amount:,.0f} 跳升至 {after_gap_amount:,.0f}，"
            f"相鄰差額為 {max_gap_amount:,.0f}，"
            f"相鄰差距比例為 {gap_ratio:.1%}。"
            f"本報表將 {before_gap_amount:,.0f} 以上客戶標註為 VIP 客戶。"
        )

        gap_table["是否斷層起始點"] = np.where(
            gap_table[money_col] == gap_threshold,
            "是",
            "否"
        )

        gap_table["是否斷層極端客戶"] = np.where(
            gap_table[money_col] >= gap_threshold,
            "是",
            "否"
        )

        gap_table["斷層法分類"] = np.where(
            gap_table[money_col] >= gap_threshold,
            "斷層極端客戶",
            "一般客戶"
        )

        gap_table_indexed = gap_table.set_index("index")

        result["金額排序"] = gap_table_indexed["金額排序"]
        result["相鄰前一筆金額"] = gap_table_indexed["相鄰前一筆金額"]
        result["相鄰差額"] = gap_table_indexed["相鄰差額"]
        result["是否斷層起始點"] = gap_table_indexed["是否斷層起始點"]
        result["是否斷層極端客戶"] = gap_table_indexed["是否斷層極端客戶"]
        result["斷層法分類"] = gap_table_indexed["斷層法分類"]
        result["斷層起始門檻"] = gap_threshold

        gap_info = {
            "gap_found": True,
            "gap_threshold": gap_threshold,
            "max_gap_amount": max_gap_amount,
            "before_gap_amount": before_gap_amount,
            "after_gap_amount": after_gap_amount,
            "gap_ratio": gap_ratio,
            "method_note": method_note
        }

        return result, gap_info, gap_table


    # 執行 Monetary 最大相鄰差額斷層法
    df, gap_info, gap_table = apply_monetary_gap_method(df, money_col="Monetary_raw")

    print("\n=== Monetary 改良式最大相鄰差額斷層法 ===")
    print(gap_info["method_note"])

    if pd.notna(gap_info["gap_threshold"]):
        print(f"斷層起始門檻：{gap_info['gap_threshold']:,.0f}")
        print(f"最大相鄰差額：{gap_info['max_gap_amount']:,.0f}")


    # =====================================
    # 3-2. M 分數計算：一般客戶五分位排名法 + VIP 倍率對數平滑法
    # =====================================

    # 設定 VIP 平滑分數參數
    # 一般金額客戶：正式 M 分數採市面常用的五分位排名法，落在 1～5 分
    # VIP 高貢獻客戶：採倍率對數平滑法，讓 VIP 之間依金額差異逐步加分
    VIP_M_BASE_SCORE = 6
    VIP_M_GROWTH_MULTIPLIER = float(vip_m_log_base)

    # 依斷層門檻建立 VIP 判斷欄位
    if pd.notna(gap_info.get("gap_threshold", np.nan)):
        df["是否VIP金額"] = np.where(
            df["Monetary_raw"] >= gap_info["gap_threshold"],
            "是",
            "否"
        )
    else:
        df["是否VIP金額"] = "否"

    # 預設分數
    # M_score：正式 M 分數
    # M_score_smooth：平滑分數，本版本主要用於 VIP 客戶；一般客戶則與正式 M 分數相同
    # M_score_group：輔助欄位，保留五分位組別
    df["M_score"] = 1
    df["M_score_smooth"] = 1.0
    df["M_score_group"] = 1

    # -------------------------------------------------
    # 一般金額客戶：五分位排名法（市面常用 RFM Monetary 評分方式）
    # -------------------------------------------------
    # 設計邏輯：
    # 1. 只針對非 VIP 且有正向銷售金額的客戶計算。
    # 2. 將一般客戶依 Monetary_raw 由小到大排名。
    # 3. 依排名平均切成 5 組，分別給 1～5 分。
    # 4. 使用 rank(method="average")，相同金額會取得相同平均排名。
    # 5. 一般客戶不額外使用平滑小數；M_score_smooth 與正式 M_score 相同。
    normal_money_mask = (
        (df["是否VIP金額"] == "否") &
        (df["Monetary_raw"] > 0)
    )

    normal_money = df.loc[normal_money_mask, "Monetary_raw"]

    if len(normal_money) > 1:
        # 市面常用五分位排名法：依累積金額排名切成五等分
        normal_rank = normal_money.rank(method="average", ascending=True)
        normal_count = len(normal_money)

        normal_group = (
            np.ceil(normal_rank / normal_count * 5)
            .astype(int)
            .clip(1, 5)
        )

        df.loc[normal_money_mask, "M_score"] = normal_group
        df.loc[normal_money_mask, "M_score_smooth"] = normal_group.astype(float)
        df.loc[normal_money_mask, "M_score_group"] = normal_group

    elif len(normal_money) == 1:
        # 若一般正向金額客戶只有一位，給中間偏高的 3 分，避免過度極端
        df.loc[normal_money_mask, "M_score"] = 3
        df.loc[normal_money_mask, "M_score_smooth"] = 3.0
        df.loc[normal_money_mask, "M_score_group"] = 3

    # -------------------------------------------------
    # VIP 高貢獻客戶：倍率對數平滑法
    # -------------------------------------------------
    # 設計邏輯：
    # 1. VIP 客戶仍由斷層法門檻判定。
    # 2. VIP 的 M_score_smooth 採對數倍率法，使 VIP 客戶之間可依金額差異拉開。
    # 3. 對數平滑可避免 VIP 客戶因金額差距過大而分數暴衝。
    #
    # 公式：
    # M平滑 = 6 + log以 VIP_M_LOG_BASE 為底(客戶金額 ÷ VIP門檻)
    vip_score_reference_money = np.nan
    vip_mask = df["是否VIP金額"] == "是"

    if vip_mask.sum() > 0 and pd.notna(gap_info.get("gap_threshold", np.nan)):
        vip_threshold = float(gap_info["gap_threshold"])

        if vip_threshold > 0:
            vip_score_reference_money = vip_threshold

            vip_score_smooth = (
                VIP_M_BASE_SCORE +
                np.log(df.loc[vip_mask, "Monetary_raw"] / vip_score_reference_money) /
                np.log(VIP_M_GROWTH_MULTIPLIER)
            )

            vip_score_smooth = vip_score_smooth.clip(lower=VIP_M_BASE_SCORE).round(2)

            # M平滑分數保留小數，只用於 VIP 客戶細部比較
            df.loc[vip_mask, "M_score_smooth"] = vip_score_smooth

            # 正式 M 分數採四捨五入後的整數
            df.loc[vip_mask, "M_score"] = (
                np.floor(vip_score_smooth + 0.5)
                .astype(int)
            )

            # 輔助群組欄位
            df.loc[vip_mask, "M_score_group"] = VIP_M_BASE_SCORE

        else:
            df.loc[vip_mask, "M_score_smooth"] = VIP_M_BASE_SCORE
            df.loc[vip_mask, "M_score"] = VIP_M_BASE_SCORE
            df.loc[vip_mask, "M_score_group"] = VIP_M_BASE_SCORE

    elif vip_mask.sum() > 0:
        df.loc[vip_mask, "M_score_smooth"] = VIP_M_BASE_SCORE
        df.loc[vip_mask, "M_score"] = VIP_M_BASE_SCORE
        df.loc[vip_mask, "M_score_group"] = VIP_M_BASE_SCORE

    # 完全沒有銷售金額者，M 分數固定為 1
    df.loc[df["Monetary_raw"] <= 0, "M_score"] = 1
    df.loc[df["Monetary_raw"] <= 0, "M_score_smooth"] = 1.0
    df.loc[df["Monetary_raw"] <= 0, "M_score_group"] = 1

    # 統一格式
    df["M_score"] = df["M_score"].astype(int)
    df["M_score_smooth"] = df["M_score_smooth"].astype(float).round(2)
    df["M_score_group"] = df["M_score_group"].astype(int)


    # 正式 RFM 總分：使用整數 R、F、M 分數
    df["RFM_total"] = (
        df["R_score"].astype(int) +
        df["F_score"].astype(int) +
        df["M_score"].astype(int)
    ).astype(int)

    # 平滑 RFM 總分：
    # 一般客戶因 M_score_smooth = M_score，故與正式總分相同；
    # VIP 客戶因 M_score_smooth 保留小數，可用於 VIP 內部細部比較
    df["RFM_smooth_total"] = (
        df["R_score"].astype(float) +
        df["F_score"].astype(float) +
        df["M_score_smooth"].astype(float)
    ).round(2)

    # 金額加權平滑總分：
    # 一般客戶使用五分位 M 分數，VIP 客戶使用平滑 M 分數
    M_WEIGHT = 2

    df["RFM_money_weighted_total"] = (
        df["R_score"].astype(float) +
        df["F_score"].astype(float) +
        df["M_score_smooth"].astype(float) * M_WEIGHT
    ).round(2)

    # RFM_score：以字串保留 R/F/M 結構
    df["RFM_score"] = (
        df["R_score"].astype(str) + "-" +
        df["F_score"].astype(str) + "-" +
        df["M_score"].astype(str)
    )


    # =====================================
    # 4. K-means 分群：排除 VIP 後分群
    # =====================================

    # VIP 客戶的 M 分數可能超過 5 分，若放入 K-means 會影響一般客戶分群，
    # 因此本報表先排除 VIP，只針對一般金額客戶做 K-means 分群。
    df["cluster"] = np.nan
    df["客群類型"] = ""

    non_vip_cluster_mask = df["是否VIP金額"] == "否"
    X_non_vip = df.loc[non_vip_cluster_mask, ["R_score", "F_score", "M_score"]].copy()

    if len(X_non_vip) >= 3:
        kmeans = KMeans(n_clusters=3, random_state=42, n_init=20)
        df.loc[non_vip_cluster_mask, "cluster"] = kmeans.fit_predict(X_non_vip)

        cluster_summary_num = df.loc[non_vip_cluster_mask].groupby("cluster")[[
            "R_score", "F_score", "M_score",
            "RFM_total", "Monetary_raw",
            "Frequency_raw", "Recency_raw"
        ]].mean().round(2).reset_index()

        high_cluster = cluster_summary_num.loc[
            cluster_summary_num["RFM_total"].idxmax(),
            "cluster"
        ]

        risk_cluster = cluster_summary_num.loc[
            cluster_summary_num["RFM_total"].idxmin(),
            "cluster"
        ]

        general_cluster = [
            c for c in cluster_summary_num["cluster"]
            if c not in [high_cluster, risk_cluster]
        ][0]

        cluster_name_map = {
            high_cluster: "高價值客戶",
            general_cluster: "一般客戶",
            risk_cluster: "有流失風險客戶"
        }

        df.loc[non_vip_cluster_mask, "客群類型"] = (
            df.loc[non_vip_cluster_mask, "cluster"].map(cluster_name_map)
        )

    else:
        # 若資料量太少，改以 RFM 總分做簡易分層
        df.loc[non_vip_cluster_mask, "客群類型"] = np.where(
            df.loc[non_vip_cluster_mask, "RFM_total"] >= 12,
            "高價值客戶",
            np.where(
                df.loc[non_vip_cluster_mask, "RFM_total"] <= 7,
                "有流失風險客戶",
                "一般客戶"
            )
        )

    # VIP 客戶獨立標示，不參與 K-means
    df.loc[df["是否VIP金額"] == "是", "客群類型"] = "VIP客戶"

    # 保險處理：避免客群類型空白
    df.loc[df["客群類型"] == "", "客群類型"] = "一般客戶"


    # =====================================
    # 5. 企業實用欄位
    # =====================================

    def priority_level(segment):
        if segment == "VIP客戶":
            return "最高"
        elif segment == "有流失風險客戶":
            return "緊急"
        elif segment == "高價值客戶":
            return "重要"
        return "一般"


    def transaction_pattern(row):
        """
        依分析月份的購買狀況判斷交易型態。
        適用 1~3、1~6、1~12 或任意月份區間。
        """
        active_flags = [row[col] > 0 for col in month_cols]
        active_count = sum(active_flags)

        if active_count == 0:
            return "無有效交易"

        if active_count == len(month_cols):
            return "連續活躍"

        if active_flags[-1]:
            if active_count >= max(2, int(np.ceil(len(month_cols) * 0.5))):
                return "近期活躍"
            return "近期新活躍"

        if not active_flags[-1] and any(active_flags[:-1]):
            return "近期下降"

        return "間歇購買"


    def immediate_followup(row):
        """
        是否需立即追蹤：
        1. 流失風險客戶：是
        2. 一般客戶且最後分析月份沒有購買：是
        3. 其餘：否
        """
        latest_month_col = month_cols[-1]

        if row["客群類型"] == "有流失風險客戶":
            return "是"

        if row["客群類型"] == "一般客戶" and row[latest_month_col] <= 0:
            return "是"

        return "否"


    df["追蹤優先級"] = df["客群類型"].apply(priority_level)
    df["交易型態"] = df.apply(transaction_pattern, axis=1)
    df["是否需立即追蹤"] = df.apply(immediate_followup, axis=1)

    # 以下欄位保留給管理者後續手動填寫
    df["追蹤備註"] = ""
    df["回訪結果"] = ""
    df["下次追蹤日期"] = ""

    segment_reason_map = {
        "VIP客戶": "銷售金額達斷層法極端值門檻，屬於公司營收貢獻明顯突出的重點客戶。",
        "高價值客戶": "近期仍有購買，活躍度高且累計銷售金額高。",
        "一般客戶": "近期仍有交易，但購買頻率或金額屬中等。",
        "有流失風險客戶": "近期購買表現偏弱，購買頻率與金額偏低。"
    }

    segment_action_map = {
        "VIP客戶": "列為最高優先維繫對象，由業務或主管定期追蹤需求、交期與報價狀況。",
        "高價值客戶": "優先維繫，定期關懷並主動提供報價與交期服務。",
        "一般客戶": "持續追蹤需求，提升購買頻率與客單價。",
        "有流失風險客戶": "優先聯繫，確認是否需求下降、轉單或沉睡。"
    }


    # =====================================
    # 6. 摘要與清單
    # =====================================

    summary = df.groupby("客群類型").agg(
        客戶數=("customer_id", "count"),
        平均R分數=("R_score", "mean"),
        平均F分數=("F_score", "mean"),
        平均M分數=("M_score", "mean"),
        平均RFM總分=("RFM_total", "mean"),
        平均銷售金額=("Monetary_raw", "mean"),
        客群總銷售金額=("Monetary_raw", "sum"),
        平均活躍月份數=("Frequency_raw", "mean"),
        平均未交易月份數=("Recency_raw", "mean")
    ).round(2).reset_index()

    if total_sales > 0:
        summary["客群銷售占比"] = summary["客群總銷售金額"] / total_sales
    else:
        summary["客群銷售占比"] = 0

    order = ["VIP客戶", "高價值客戶", "一般客戶", "有流失風險客戶"]

    summary["sort_key"] = summary["客群類型"].map({
        k: i for i, k in enumerate(order)
    })

    summary = summary.sort_values("sort_key").drop(columns="sort_key")

    # 確保聯絡欄位存在
    for col in ["聯絡人", "聯絡電話", "聯絡人手機"]:
        if col not in df.columns:
            df[col] = ""

        df[col] = df[col].fillna("").astype(str).str.strip()

    vip_value = df[df["客群類型"] == "VIP客戶"].sort_values(
        ["Monetary_raw", "RFM_money_weighted_total"],
        ascending=[False, False]
    ).copy()

    high_value = df[df["客群類型"] == "高價值客戶"].sort_values(
        ["RFM_money_weighted_total", "Monetary_raw"],
        ascending=[False, False]
    ).copy()

    general = df[df["客群類型"] == "一般客戶"].sort_values(
        ["RFM_money_weighted_total", "Monetary_raw"],
        ascending=[False, False]
    ).copy()

    risk = df[df["客群類型"] == "有流失風險客戶"].sort_values(
        ["R_score", "F_score", "Monetary_raw"],
        ascending=[True, True, True]
    ).copy()


    # =====================================
    # 7. 圖表
    # =====================================

    # 圖1：K-means 客群分布結果
    counts = df["客群類型"].value_counts().reindex(order).fillna(0)

    plt.figure(figsize=(8, 5))
    bars = plt.bar(counts.index, counts.values)
    plt.title("K-means 客群分布結果")
    plt.xlabel("客群類型")
    plt.ylabel("客戶數")

    for bar, val in zip(bars, counts.values):
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            val,
            f"{int(val)}",
            ha="center",
            va="bottom"
        )

    plt.tight_layout()
    plt.savefig(count_img, dpi=180, bbox_inches="tight")
    plt.close()


    # 圖2：各客群平均 RFM 分數比較
    score_summary = summary.set_index("客群類型")[[
        "平均R分數", "平均F分數", "平均M分數"
    ]].reindex(order).fillna(0)

    x = np.arange(len(score_summary.index))
    width = 0.22

    plt.figure(figsize=(9, 5))

    for i, col in enumerate(score_summary.columns):
        bars = plt.bar(
            x + (i - 1) * width,
            score_summary[col].values,
            width=width,
            label=col
        )

        for bar, val in zip(bars, score_summary[col].values):
            plt.text(
                bar.get_x() + bar.get_width() / 2,
                val,
                f"{val:.2f}",
                ha="center",
                va="bottom",
                fontsize=8
            )

    plt.xticks(x, score_summary.index)
    plt.ylabel("平均分數")
    plt.title("各客群平均 RFM 分數比較")
    plt.legend()
    plt.tight_layout()
    plt.savefig(rfm_img, dpi=180, bbox_inches="tight")
    plt.close()


    # 圖3：各客群平均銷售金額比較
    money_summary = summary.set_index("客群類型")["平均銷售金額"].reindex(order).fillna(0)

    plt.figure(figsize=(8, 5))
    bars = plt.bar(money_summary.index, money_summary.values)
    plt.title("各客群平均銷售金額比較")
    plt.xlabel("客群類型")
    plt.ylabel("平均銷售金額")

    for bar, val in zip(bars, money_summary.values):
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            val,
            f"{val:,.0f}",
            ha="center",
            va="bottom"
        )

    plt.tight_layout()
    plt.savefig(money_img, dpi=180, bbox_inches="tight")
    plt.close()


    # 圖4：各客群銷售貢獻度比較
    contrib_summary = summary.set_index("客群類型")["客群銷售占比"].reindex(order).fillna(0)

    plt.figure(figsize=(8, 5))
    bars = plt.bar(contrib_summary.index, contrib_summary.values)
    plt.title("各客群銷售貢獻度比較")
    plt.xlabel("客群類型")
    plt.ylabel("銷售占比")

    for bar, val in zip(bars, contrib_summary.values):
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            val,
            f"{val:.1%}",
            ha="center",
            va="bottom"
        )

    plt.tight_layout()
    plt.savefig(contrib_img, dpi=180, bbox_inches="tight")
    plt.close()


    # 圖5：Monetary 最大相鄰差額斷層法曲線圖
    gap_plot = gap_table.copy().sort_values("金額排序")

    plt.figure(figsize=(10, 5))
    plt.plot(
        gap_plot["金額排序"],
        gap_plot["Monetary_raw"],
        marker="o",
        linewidth=1.5,
        markersize=3,
        label="客戶金額排序曲線"
    )

    if gap_info["gap_found"] and pd.notna(gap_info["gap_threshold"]):
        gap_start = gap_plot[
            gap_plot["Monetary_raw"] == gap_info["gap_threshold"]
        ].head(1)

        if not gap_start.empty:
            x0 = gap_start["金額排序"].iloc[0]
            y0 = gap_start["Monetary_raw"].iloc[0]

            plt.scatter([x0], [y0], s=120, zorder=5, label="VIP起始點")
            plt.text(
                x0,
                y0,
                f"  VIP起始點\n  {y0:,.0f}",
                va="bottom",
                fontsize=9
            )

        extreme_points = gap_plot[
            gap_plot["Monetary_raw"] >= gap_info["gap_threshold"]
        ]

        plt.scatter(
            extreme_points["金額排序"],
            extreme_points["Monetary_raw"],
            s=70,
            zorder=5,
            label="VIP客戶"
        )

    plt.title("Monetary 改良式最大相鄰差額斷層法")
    plt.xlabel("金額由小到大排序序號")
    plt.ylabel("分析期間合計銷售金額")
    plt.legend()
    plt.grid(axis="y", linestyle="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(gap_img, dpi=180, bbox_inches="tight")
    plt.close()


    # =====================================
    # 8. Excel 樣式
    # =====================================

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)

    fill_high = PatternFill("solid", fgColor="E2F0D9")
    fill_general = PatternFill("solid", fgColor="FFF2CC")
    fill_risk = PatternFill("solid", fgColor="F4CCCC")
    fill_extreme = PatternFill("solid", fgColor="FCE4D6")

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)


    def get_fill_by_segment(segment_name):
        if segment_name == "VIP客戶":
            return fill_extreme
        elif segment_name == "高價值客戶":
            return fill_high
        elif segment_name == "一般客戶":
            return fill_general
        elif segment_name == "有流失風險客戶":
            return fill_risk

        return None


    def set_header_style(ws, header_row=1):
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border


    def color_rows_by_segment(ws, segment_col_idx, start_row, end_row, end_col):
        for row in range(start_row, end_row + 1):
            segment_value = ws.cell(row=row, column=segment_col_idx).value
            fill = get_fill_by_segment(segment_value)

            if fill:
                for col in range(1, end_col + 1):
                    ws.cell(row=row, column=col).fill = fill
                    ws.cell(row=row, column=col).border = border


    def write_fixed_info_block(ws, start_row=1, end_col=20):
        ws[f"A{start_row}"] = "客群追蹤說明"
        ws[f"A{start_row}"].font = title_font

        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=end_col
        )

        ws.row_dimensions[start_row].height = 18

        header_row = start_row + 1

        ws.cell(row=header_row, column=1, value="客群類型")
        ws.cell(row=header_row, column=2, value="追蹤優先級")
        ws.cell(row=header_row, column=3, value="分群原因")
        ws.cell(row=header_row, column=6, value="建議追蹤方式")

        ws.merge_cells(
            start_row=header_row,
            start_column=3,
            end_row=header_row,
            end_column=5
        )

        ws.merge_cells(
            start_row=header_row,
            start_column=6,
            end_row=header_row,
            end_column=9
        )

        for col in range(1, 10):
            ws.cell(row=header_row, column=col).fill = header_fill
            ws.cell(row=header_row, column=col).font = header_font
            ws.cell(row=header_row, column=col).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            ws.cell(row=header_row, column=col).border = border

        ws.row_dimensions[header_row].height = 18

        current = header_row + 1

        for seg in order:
            ws.cell(row=current, column=1, value=seg)
            ws.cell(row=current, column=2, value=priority_level(seg))
            ws.cell(row=current, column=3, value=segment_reason_map[seg])
            ws.cell(row=current, column=6, value=segment_action_map[seg])

            ws.merge_cells(
                start_row=current,
                start_column=3,
                end_row=current,
                end_column=5
            )

            ws.merge_cells(
                start_row=current,
                start_column=6,
                end_row=current,
                end_column=9
            )

            fill = get_fill_by_segment(seg)

            for col in range(1, 10):
                ws.cell(row=current, column=col).fill = fill
                ws.cell(row=current, column=col).border = border
                ws.cell(row=current, column=col).alignment = Alignment(
                    horizontal="left" if col in [3, 6] else "center",
                    vertical="center",
                    wrap_text=False
                )

            ws.row_dimensions[current].height = 18
            current += 1

        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws.column_dimensions[col].width = 16

        return current


    # =====================================
    # 9. 群組摘要
    # =====================================

    ws1 = wb.active
    ws1.title = "群組摘要"

    ws1["A1"] = "RFM + K-means 客群分析摘要"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:K1")
    ws1.row_dimensions[1].height = 18

    for r in dataframe_to_rows(summary, index=False, header=True):
        ws1.append(r)

    set_header_style(ws1, 2)

    summary_widths = {
        "A": 18,
        "B": 12,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 16,
        "G": 16,
        "H": 18,
        "I": 16,
        "J": 20,
        "K": 16
    }

    for col, width in summary_widths.items():
        ws1.column_dimensions[col].width = width

    for row in range(3, ws1.max_row + 1):
        seg = ws1.cell(row=row, column=1).value
        fill = get_fill_by_segment(seg)

        if fill:
            for col in range(1, ws1.max_column + 1):
                ws1.cell(row=row, column=col).fill = fill
                ws1.cell(row=row, column=col).border = border

        for col in [7, 8]:
            ws1.cell(row=row, column=col).number_format = '#,##0'

        ws1.cell(row=row, column=11).number_format = '0.0%'
        ws1.row_dimensions[row].height = 18


    # =====================================
    # 9-1. RFM 評分規範
    # =====================================

    ws_rule = wb.create_sheet("RFM評分規範")

    ws_rule["A1"] = "RFM 評分規範"
    ws_rule["A1"].font = title_font
    ws_rule.merge_cells("A1:D1")

    rule_rows = [
        ["項目", "評分方式", "分數範圍", "說明"],
        [
            "R分數 Recency",
            "依最近一次交易距離最後分析月份的間隔換算",
            "1～5分",
            "距離最後分析月份越近，分數越高；最後月份有交易者通常為5分，分析期間完全無交易給1分。"
        ],
        [
            "F分數 Frequency",
            "有效交易月份數 ÷ 分析月份數 × 5，採向上取整",
            "1～5分",
            "交易月份占比越高，代表購買穩定度越高，分數越高。"
        ],
        [
            "M分數 Monetary",
            "一般客戶採五分位排名法；VIP客戶採倍率對數平滑法",
            "一般客戶1～5分；VIP客戶6分以上",
            "一般金額客戶採市面常用的五分位排名法，先計算分析期間累積銷售金額，再依金額由小到大排序，依排名平均切成五組並給予1～5分。一般客戶不額外使用平滑小數，M平滑分數與正式M分數相同；VIP客戶則以VIP門檻為基準，採倍率對數平滑法計算6分以上之M平滑分數，正式M分數由M平滑分數四捨五入取得。"
        ],
        [
            "RFM總分",
            "R分數 + F分數 + M分數",
            "最低3分，最高會隨VIP金額增加",
            "RFM總分採R、F、M正式分數加總，以整數呈現，方便管理者快速判讀；一般客戶的M平滑分數與正式M分數相同，VIP客戶才保留小數平滑分數，用於VIP內部或同分客戶之間的細部排序與比較。"
        ],
        [
            "分析月份",
            ", ".join(month_display_cols),
            f"{analysis_month_count}個月",
            "本報表依使用者設定的 ANALYSIS_MONTHS 進行分析，並以所選月份區間的最後月份作為R分數基準。"
        ],
    ]

    for r_idx, row in enumerate(rule_rows, start=3):
        for c_idx, value in enumerate(row, start=1):
            ws_rule.cell(row=r_idx, column=c_idx, value=value)

    set_header_style(ws_rule, 3)

    for row in range(4, 9):
        for col in range(1, 5):
            ws_rule.cell(row=row, column=col).border = border
            ws_rule.cell(row=row, column=col).alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

    ws_rule.column_dimensions["A"].width = 22
    ws_rule.column_dimensions["B"].width = 46
    ws_rule.column_dimensions["C"].width = 24
    ws_rule.column_dimensions["D"].width = 78


    # =====================================
    # 9-2. 本次 R / F / M 分數判定表
    # =====================================

    # ---------- 本次 R 分數判定表 ----------
    r_table_start = 11

    ws_rule[f"A{r_table_start}"] = "本次 R 分數判定表"
    ws_rule[f"A{r_table_start}"].font = title_font
    ws_rule.merge_cells(start_row=r_table_start, start_column=1, end_row=r_table_start, end_column=4)

    r_header_row = r_table_start + 1
    r_headers = ["最近交易情況", "距離最後分析月份", "R分數", "管理意義"]

    for col_idx, header in enumerate(r_headers, start=1):
        cell = ws_rule.cell(row=r_header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    r_data_start = r_header_row + 1

    for distance in range(analysis_month_count):
        score = calc_r_score_dynamic(distance, analysis_month_count)
        month_name = month_display_cols[-1 - distance]

        if distance == 0:
            meaning = "最後分析月份仍有交易，近期往來最活躍。"
        elif score >= 4:
            meaning = "距離最後分析月份很近，近期仍有交易跡象。"
        elif score == 3:
            meaning = "分析期間中段曾有交易，近期活躍度普通。"
        elif score == 2:
            meaning = "距離最後分析月份較久，近期交易偏弱。"
        else:
            meaning = "距離最後分析月份最久，近期交易明顯偏弱。"

        row_data = [f"最後交易月份：{month_name}", f"{distance} 個月", score, meaning]
        current_row = r_data_start + distance

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_rule.cell(row=current_row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in [2, 3] else "left",
                vertical="center",
                wrap_text=True
            )

    no_trade_row = r_data_start + analysis_month_count
    no_trade_data = ["分析期間完全無交易", "無有效交易月份", 1, "分析期間內沒有任何正向交易，給予最低R分數。"]

    for col_idx, value in enumerate(no_trade_data, start=1):
        cell = ws_rule.cell(row=no_trade_row, column=col_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col_idx in [2, 3] else "left",
            vertical="center",
            wrap_text=True
        )


    # ---------- 本次 F 分數判定表 ----------
    f_table_start = no_trade_row + 3

    ws_rule[f"A{f_table_start}"] = "本次 F 分數判定表"
    ws_rule[f"A{f_table_start}"].font = title_font
    ws_rule.merge_cells(start_row=f_table_start, start_column=1, end_row=f_table_start, end_column=4)

    f_header_row = f_table_start + 1
    f_headers = ["有效交易月份數", "計算方式", "F分數", "管理意義"]

    for col_idx, header in enumerate(f_headers, start=1):
        cell = ws_rule.cell(row=f_header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    f_data_start = f_header_row + 1

    for active_months in range(0, analysis_month_count + 1):
        score = calc_f_score_dynamic(active_months, analysis_month_count)

        if active_months == 0:
            formula_text = "無交易，直接給1分"
            meaning = "分析期間內沒有交易，購買穩定度最低。"
        else:
            formula_text = f"向上取整（{active_months} ÷ {analysis_month_count} × 5）"

            if score >= 5:
                meaning = "分析期間內交易月份占比最高，購買穩定度佳。"
            elif score >= 4:
                meaning = "交易月份占比較高，購買穩定度良好。"
            elif score == 3:
                meaning = "交易月份占比中等，購買穩定度普通。"
            elif score == 2:
                meaning = "交易月份占比較低，購買穩定度偏弱。"
            else:
                meaning = "交易月份占比很低，購買穩定度低。"

        row_data = [f"{active_months} 個月有交易", formula_text, score, meaning]
        current_row = f_data_start + active_months

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_rule.cell(row=current_row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in [1, 3] else "left",
                vertical="center",
                wrap_text=True
            )


    # ---------- 本次 M 分數判定表 ----------
    m_table_start = f_data_start + analysis_month_count + 4

    ws_rule[f"A{m_table_start}"] = "本次 M 分數判定表"
    ws_rule[f"A{m_table_start}"].font = title_font
    ws_rule.merge_cells(start_row=m_table_start, start_column=1, end_row=m_table_start, end_column=4)

    m_header_row = m_table_start + 1
    m_headers = ["評分對象", "本組實際金額範圍", "M分數", "計算依據／說明"]

    for col_idx, header in enumerate(m_headers, start=1):
        cell = ws_rule.cell(row=m_header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


    def format_money_range(min_value, max_value):
        """將金額範圍轉成報表顯示文字。"""
        if pd.isna(min_value) or pd.isna(max_value):
            return "本次無客戶"

        min_value = float(min_value)
        max_value = float(max_value)

        if min_value == max_value:
            return f"{min_value:,.0f}"

        return f"{min_value:,.0f} ～ {max_value:,.0f}"


    def build_normal_m_score_range_rows(normal_df):
        """
        產生一般客戶 M 分數 1～5 分的實際金額範圍說明。
        說明：
        1. 一般客戶 M 分數採市面常用五分位排名法。
        2. 先依分析期間累積銷售金額由小到大排序。
        3. 再依排名平均切成五組，分別給 1～5 分。
        4. 表列金額範圍為本次資料中該分數組的實際最低～最高金額。
        """
        rows = []

        if normal_df.empty:
            return rows

        temp = normal_df.copy()
        temp = temp[temp["Monetary_raw"] > 0].copy()

        if temp.empty:
            return rows

        temp = temp.sort_values("Monetary_raw", ascending=True).copy()
        temp["_M_rank_order"] = np.arange(1, len(temp) + 1)

        normal_count = len(temp)

        for score in range(1, 6):
            group_df = temp[temp["M_score"] == score].copy()

            if group_df.empty:
                range_text = "本次無客戶"
                desc_text = (
                    f"公式：M分數＝向上取整（金額排名 ÷ {normal_count} × 5）。"
                    f"本次資料中沒有客戶落在 M={score} 分。"
                )
            else:
                min_amount = group_df["Monetary_raw"].min()
                max_amount = group_df["Monetary_raw"].max()
                group_count = len(group_df)

                rank_min = int(group_df["_M_rank_order"].min())
                rank_max = int(group_df["_M_rank_order"].max())

                range_text = format_money_range(min_amount, max_amount)

                desc_text = (
                    f"公式：M分數＝向上取整（金額排名 ÷ {normal_count} × 5）。"
                    f"本組為第{rank_min}～{rank_max}名，共{group_count}位；"
                    f"表列金額為本次資料中該組實際最低～最高金額，並非固定連續門檻。"
                )

            rows.append([
                f"一般金額客戶－{score}分",
                range_text,
                f"{score}分",
                desc_text
            ])

        return rows


    current_row = m_header_row + 1

    normal_m_range_df = df[(df["是否VIP金額"] == "否") & (df["Monetary_raw"] > 0)].copy()
    normal_score_rows = build_normal_m_score_range_rows(normal_m_range_df)

    if normal_score_rows:
        for normal_row_data in normal_score_rows:
            for col_idx, value in enumerate(normal_row_data, start=1):
                cell = ws_rule.cell(row=current_row, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center" if col_idx in [1, 3] else "left",
                    vertical="center",
                    wrap_text=True
                )

            current_row += 1
    else:
        normal_row_data = [
            "一般金額客戶",
            "本次無一般金額客戶",
            "無",
            "本次無一般金額客戶可計算五分位排名法 M 分數。"
        ]

        for col_idx, value in enumerate(normal_row_data, start=1):
            cell = ws_rule.cell(row=current_row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in [1, 3] else "left",
                vertical="center",
                wrap_text=True
            )

        current_row += 1

    vip_customer_df = df[df["是否VIP金額"] == "是"].copy()

    if not vip_customer_df.empty:
        vip_count = len(vip_customer_df)
        vip_range_text = f"{gap_info['gap_threshold']:,.0f} 以上"

        vip_formula_text = (
            f"VIP客戶共{vip_count}位；"
            f"VIP判定門檻為{gap_info['gap_threshold']:,.0f}以上；"
            f"M平滑＝{VIP_M_BASE_SCORE}+log(金額÷{vip_score_reference_money:,.0f})"
            f"÷log({VIP_M_GROWTH_MULTIPLIER})；"
            f"其中{vip_score_reference_money:,.0f}為VIP分數計算基準，"
            f"用於讓VIP客戶依金額倍率逐步加分，並避免分數差距過大；"
            f"正式M分數由M平滑分數四捨五入取得。"
        )
    else:
        vip_range_text = "本次無VIP高貢獻客戶"
        vip_formula_text = "本次無客戶達到VIP判定門檻。"

    vip_row_data = [
        "VIP高貢獻客戶",
        vip_range_text,
        "M平滑：6分以上；M分數：6分以上整數",
        vip_formula_text
    ]

    for col_idx, value in enumerate(vip_row_data, start=1):
        cell = ws_rule.cell(row=current_row, column=col_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col_idx in [1, 3] else "left",
            vertical="center",
            wrap_text=True
        )
        cell.fill = fill_extreme

    current_row += 1

    for row in range(1, ws_rule.max_row + 1):
        ws_rule.row_dimensions[row].height = 36


    # =====================================
    # 10. 圖表總覽
    # =====================================

    ws_chart = wb.create_sheet("圖表總覽")

    ws_chart["A1"] = "客群分析圖表總覽"
    ws_chart["A1"].font = title_font

    for col in [
        "A", "B", "C", "D", "E", "F", "G", "H", "I",
        "J", "K", "L", "M", "N", "O", "P", "Q",
        "R", "S", "T", "U", "V", "W", "X", "Y", "Z"
    ]:
        ws_chart.column_dimensions[col].width = 14

    for r in range(1, 160):
        ws_chart.row_dimensions[r].height = 22

    chart_items = [
        ("圖1：K-means 客群分布結果", count_img, 720, 520),
        ("圖2：各客群平均 RFM 分數比較", rfm_img, 720, 520),
        ("圖3：各客群平均銷售金額比較", money_img, 720, 520),
        ("圖4：各客群銷售貢獻度比較", contrib_img, 720, 520),
        ("圖5：Monetary 改良式最大相鄰差額斷層法", gap_img, 720, 520),
    ]

    charts_per_row = 3
    col_positions = ["A", "J", "S"]
    row_start = 3
    row_step = 24

    for idx, (title, img_path, img_width, img_height) in enumerate(chart_items):
        row_group = idx // charts_per_row
        col_group = idx % charts_per_row

        start_col = col_positions[col_group]
        title_row = row_start + row_group * row_step
        image_row = title_row + 1

        ws_chart[f"{start_col}{title_row}"] = title
        ws_chart[f"{start_col}{title_row}"].font = bold_font

        img = XLImage(str(img_path))
        img.width = img_width
        img.height = img_height

        ws_chart.add_image(img, f"{start_col}{image_row}")


    # =====================================
    # 11. 客戶分群結果
    # =====================================

    ws2 = wb.create_sheet("客戶分群結果")

    # 欄位數：
    # 客戶代號、客戶名稱 + 月份欄位 + Monetary + 銷售貢獻度
    # + R/F/M/RFM + 客群/追蹤欄位 + 聯絡欄位 + 備註欄位
    main_output_col_count = 2 + len(month_cols) + 1 + 1 + 5 + 4 + 3 + 3

    next_row = write_fixed_info_block(ws2, start_row=1, end_col=main_output_col_count)
    data_start_row = next_row + 1

    ws2[f"A{data_start_row - 1}"] = "客戶分群明細"
    ws2[f"A{data_start_row - 1}"].font = bold_font
    ws2.row_dimensions[data_start_row - 1].height = 18

    cols = (
        ["customer_id", "customer_name"] +
        month_cols +
        [
            "Monetary_raw",
            "銷售貢獻度",
            "R_score", "F_score", "M_score", "M_score_smooth", "RFM_total",
            "客群類型", "追蹤優先級", "交易型態", "是否需立即追蹤",
            "聯絡人", "聯絡電話", "聯絡人手機",
            "追蹤備註", "回訪結果", "下次追蹤日期"
        ]
    )

    out_df = df[cols].copy()

    out_df["排序用客戶代號"] = pd.to_numeric(
        out_df["customer_id"],
        errors="coerce"
    )

    out_df = out_df.sort_values(
        by=["排序用客戶代號", "customer_id"],
        ascending=[True, True]
    ).drop(columns=["排序用客戶代號"]).reset_index(drop=True)

    out_df.columns = (
        ["客戶代號", "客戶名稱"] +
        month_display_cols +
        [
            "Monetary原始值",
            "銷售貢獻度",
            "R分數", "F分數", "M分數", "M平滑分數", "RFM總分",
            "客群類型", "追蹤優先級", "交易型態", "是否需立即追蹤",
            "聯絡人", "聯絡電話", "聯絡人手機",
            "追蹤備註", "回訪結果", "下次追蹤日期"
        ]
    )

    for r_idx, row in enumerate(
        dataframe_to_rows(out_df, index=False, header=True),
        start=data_start_row
    ):
        for c_idx, value in enumerate(row, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    set_header_style(ws2, data_start_row)

    # 動態欄位位置
    month_start_col = 3
    month_end_col = 2 + len(month_cols)
    monetary_col = month_end_col + 1
    contrib_col = monetary_col + 1
    r_col = contrib_col + 1
    f_col = r_col + 1
    m_col = f_col + 1
    smooth_col = m_col + 1
    rfm_total_col = smooth_col + 1
    segment_col = rfm_total_col + 1
    priority_col = segment_col + 1
    pattern_col = priority_col + 1
    follow_col = pattern_col + 1
    contact_person_col = follow_col + 1
    contact_phone_col = contact_person_col + 1
    contact_mobile_col = contact_phone_col + 1
    note_col = contact_mobile_col + 1
    result_col = note_col + 1
    next_date_col = result_col + 1

    widths = []
    widths += [16, 16]
    widths += [14] * len(month_cols)
    widths += [16, 12]
    widths += [10, 10, 12, 12, 12]
    widths += [16, 12, 14, 14]
    widths += [14, 16, 16]
    widths += [18, 18, 16]

    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    color_rows_by_segment(
        ws2,
        segment_col_idx=segment_col,
        start_row=data_start_row + 1,
        end_row=ws2.max_row,
        end_col=main_output_col_count
    )

    for row in range(data_start_row + 1, ws2.max_row + 1):
        ws2.row_dimensions[row].height = 18

        # 月份金額欄位
        for col in range(month_start_col, month_end_col + 1):
            ws2.cell(row=row, column=col).number_format = '#,##0'
            ws2.cell(row=row, column=col).border = border

        # Monetary 金額
        ws2.cell(row=row, column=monetary_col).number_format = '#,##0'
        ws2.cell(row=row, column=monetary_col).border = border

        # 銷售貢獻度
        ws2.cell(row=row, column=contrib_col).number_format = '0.0%'
        ws2.cell(row=row, column=contrib_col).border = border
        ws2.cell(row=row, column=contrib_col).alignment = Alignment(horizontal="center", vertical="center")

        # M分數與RFM總分為整數；M平滑分數保留兩位小數
        ws2.cell(row=row, column=m_col).number_format = '0'
        ws2.cell(row=row, column=smooth_col).number_format = '0.00'
        ws2.cell(row=row, column=rfm_total_col).number_format = '0'
    

        # R/F/M/RFM、客群、追蹤判斷
        for col in [
            r_col, f_col, m_col, smooth_col, rfm_total_col,
            segment_col, priority_col, pattern_col, follow_col
        ]:
            ws2.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws2.cell(row=row, column=col).border = border

        # 聯絡欄位
        for col in [contact_person_col, contact_phone_col, contact_mobile_col]:
            ws2.cell(row=row, column=col).number_format = '@'
            ws2.cell(row=row, column=col).border = border
            ws2.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # 追蹤欄位
        for col in [note_col, result_col, next_date_col]:
            ws2.cell(row=row, column=col).border = border

        # VIP客戶加強底色
        if ws2.cell(row=row, column=segment_col).value == "VIP客戶":
            for col in range(1, main_output_col_count + 1):
                ws2.cell(row=row, column=col).fill = fill_extreme

    ws2.freeze_panes = f"A{data_start_row + 1}"
    ws2.auto_filter.ref = f"A{data_start_row}:{get_column_letter(main_output_col_count)}{ws2.max_row}"


    # =====================================
    # 13. 企業實用清單
    # =====================================

    def add_action_sheet(sheet_name, source_df, segment_name):
        ws = wb.create_sheet(sheet_name)

        action_col_count = 2 + len(month_cols) + 1 + 1 + 3 + 4 + 3 + 3

        next_row = write_fixed_info_block(ws, start_row=1, end_col=action_col_count)

        title_row = next_row + 1
        ws[f"A{title_row}"] = sheet_name
        ws[f"A{title_row}"].font = bold_font

        data_start = title_row + 1

        keep = (
            ["customer_id", "customer_name"] +
            month_cols +
            [
                "Monetary_raw", "銷售貢獻度",
                "M_score", "M_score_smooth", "RFM_total", "客群類型", "追蹤優先級", "交易型態", "是否需立即追蹤",
                "聯絡人", "聯絡電話", "聯絡人手機",
                "追蹤備註", "回訪結果", "下次追蹤日期"
            ]
        )

        temp = source_df[keep].copy()

        temp.columns = (
            ["客戶代號", "客戶名稱"] +
            month_display_cols +
            [
                "合計金額", "銷售貢獻度",
                "M分數", "M平滑分數", "RFM總分",    
                "客群類型", "追蹤優先級", "交易型態", "是否需立即追蹤",
                "聯絡人", "聯絡電話", "聯絡人手機",
                "追蹤備註", "回訪結果", "下次追蹤日期"
            ]
        )

        for r_idx, row in enumerate(
            dataframe_to_rows(temp, index=False, header=True),
            start=data_start
        ):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        set_header_style(ws, data_start)

        widths = []
        widths += [12, 18]
        widths += [12] * len(month_cols)
        widths += [14, 12, 12, 12, 12, 14, 14, 14, 14]
        widths += [14, 16, 16]
        widths += [18, 16, 16]

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 動態欄位位置
        action_month_start_col = 3
        action_month_end_col = 2 + len(month_cols)
        action_monetary_col = action_month_end_col + 1
        action_contrib_col = action_monetary_col + 1
        action_m_col = action_contrib_col + 1
        action_smooth_col = action_m_col + 1
        action_rfm_total_col = action_smooth_col + 1
        action_segment_col = action_rfm_total_col + 1
        action_priority_col = action_segment_col + 1
        action_pattern_col = action_priority_col + 1
        action_follow_col = action_pattern_col + 1
        action_contact_person_col = action_follow_col + 1
        action_contact_phone_col = action_contact_person_col + 1
        action_contact_mobile_col = action_contact_phone_col + 1
        action_note_col = action_contact_mobile_col + 1
        action_result_col = action_note_col + 1
        action_next_date_col = action_result_col + 1

        fill = get_fill_by_segment(segment_name)

        for row in range(data_start + 1, ws.max_row + 1):
            for col in range(1, action_col_count + 1):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).border = border

            # 月份金額
            for col in range(action_month_start_col, action_month_end_col + 1):
                ws.cell(row=row, column=col).number_format = '#,##0'

            # 合計金額
            ws.cell(row=row, column=action_monetary_col).number_format = '#,##0'

            # 銷售貢獻度
            ws.cell(row=row, column=action_contrib_col).number_format = '0.0%'
            ws.cell(row=row, column=action_contrib_col).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row=row, column=action_m_col).number_format = '0'
            ws.cell(row=row, column=action_smooth_col).number_format = '0.00'
            ws.cell(row=row, column=action_rfm_total_col).number_format = '0'
        

            # 分類與追蹤欄位
            for col in [
                action_m_col,
                action_smooth_col,
                action_rfm_total_col,
                action_segment_col,
                action_priority_col,
                action_pattern_col,
                action_follow_col
            ]:
        
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

            # 聯絡欄位
            for col in [
                action_contact_person_col,
                action_contact_phone_col,
                action_contact_mobile_col
            ]:
                ws.cell(row=row, column=col).number_format = '@'
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

            # VIP客戶加強底色
            if ws.cell(row=row, column=action_segment_col).value == "VIP客戶":
                for col in range(1, action_col_count + 1):
                    ws.cell(row=row, column=col).fill = fill_extreme

        ws.freeze_panes = f"A{data_start + 1}"
        ws.auto_filter.ref = f"A{data_start}:{get_column_letter(action_col_count)}{ws.max_row}"


    add_action_sheet("VIP客戶維護清單", vip_value, "VIP客戶")
    add_action_sheet("高價值客戶維護清單", high_value, "高價值客戶")
    add_action_sheet("一般客戶經營清單", general, "一般客戶")
    add_action_sheet("流失風險客戶追蹤清單", risk, "有流失風險客戶")


    # =====================================
    # 14. 儲存
    # =====================================

    wb.save(output_file)

    # 刪除暫存圖片
    for tmp in [count_img, rfm_img, money_img, contrib_img, gap_img]:
        if tmp.exists():
            try:
                os.remove(tmp)
            except Exception:
                pass

    return output_file
#

# =====================================
# Streamlit 網頁介面
# =====================================

st.set_page_config(
    page_title="客群分析報表",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown(
    """
    <style>
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
        max-width: 1180px;
    }

    .app-hero {
        background: linear-gradient(135deg, #0f3d5e 0%, #1f6f8b 52%, #2f9e8f 100%);
        padding: 30px 34px;
        border-radius: 24px;
        color: white;
        box-shadow: 0 10px 28px rgba(15, 61, 94, 0.22);
        margin-bottom: 22px;
    }

    .app-hero h1 {
        margin: 0;
        font-size: 34px;
        font-weight: 800;
        letter-spacing: 0.5px;
    }

    .app-hero p {
        margin-top: 10px;
        font-size: 17px;
        line-height: 1.7;
        opacity: 0.96;
    }

    .info-card {
        background: #ffffff;
        border: 1px solid #e6edf3;
        border-radius: 18px;
        padding: 18px 20px;
        box-shadow: 0 4px 16px rgba(31, 78, 121, 0.08);
        min-height: 138px;
    }

    .info-card h3 {
        font-size: 18px;
        margin: 0 0 8px 0;
        color: #1f4e78;
        font-weight: 800;
    }

    .info-card p {
        font-size: 14px;
        line-height: 1.65;
        color: #4b5563;
        margin: 0;
    }

    .section-title {
        font-size: 22px;
        font-weight: 800;
        color: #1f4e78;
        margin-top: 28px;
        margin-bottom: 12px;
    }

    .small-note {
        color: #6b7280;
        font-size: 13px;
        line-height: 1.6;
    }

    div.stButton > button:first-child {
        height: 52px;
        border-radius: 14px;
        font-size: 17px;
        font-weight: 800;
    }

    div[data-testid="stDownloadButton"] > button {
        height: 52px;
        border-radius: 14px;
        font-size: 17px;
        font-weight: 800;
    }
    </style>
    """,
    unsafe_allow_html=True
)



# =====================================
# 系統預設設定
# =====================================
# 這兩個參數不顯示在前台，避免企業使用者操作過於複雜。
DEFAULT_VIP_GAP_RATIO = 0.5
DEFAULT_VIP_LOG_BASE = 2.5


st.markdown(
    """
    <div class="app-hero">
        <h1>📊 客群分析報表</h1>
        <p>
            上傳銷貨資料 Excel，選擇起始月份與結束月份後，系統會自動完成
            RFM 評分、VIP 斷層判定、K-means 分群、客戶追蹤清單與 Excel 報表輸出。
        </p>
    </div>
    """,
    unsafe_allow_html=True
)

card1, card2, card3 = st.columns(3)
with card1:
    st.markdown(
        """
        <div class="info-card">
            <h3>① 上傳銷貨資料</h3>
            <p>上傳銷貨資料 Excel，欄位至少需包含客戶、客戶簡稱與月份金額欄位。</p>
        </div>
        """,
        unsafe_allow_html=True
    )

with card2:
    st.markdown(
        """
        <div class="info-card">
            <h3>② 選擇分析期間</h3>
            <p>使用起始月份與結束月份設定分析範圍，不需要逐一點選月份，操作更直覺。</p>
        </div>
        """,
        unsafe_allow_html=True
    )

with card3:
    st.markdown(
        """
        <div class="info-card">
            <h3>③ 下載分析報表</h3>
            <p>分析完成後可直接下載 Excel，內含摘要、評分規範、圖表與客戶追蹤清單。</p>
        </div>
        """,
        unsafe_allow_html=True
    )

# =====================================
# 上傳資料區
# =====================================

st.markdown("### 📂 資料上傳")

sales_file = st.file_uploader(
    "上傳銷貨資料 Excel（必填）",
    type=["xlsx"],
    help="請上傳包含客戶、客戶簡稱、1月、2月、3月...等欄位的銷貨資料。"
)

customer_file = st.file_uploader(
    "上傳客戶資料統計 Excel（選填）",
    type=["xlsx"],
    help="若上傳客戶資料統計檔，報表會嘗試帶入聯絡人、聯絡電話、聯絡人手機。"
)

if customer_file is not None:
    customer_source_text = f"使用本次上傳的客戶資料：{customer_file.name}"
    st.info("本次分析將使用你上傳的客戶資料統計檔，並嘗試帶入聯絡資訊。")
else:
    customer_source_text = "未使用客戶資料統計檔，報表不會帶入聯絡資訊"
    st.info("本次未上傳客戶資料統計檔，報表中的聯絡資料欄位將保留空白。")


with st.expander("📌 銷貨資料欄位格式說明", expanded=False):
    st.markdown(
        """
        銷貨資料至少需要包含下列欄位：

        | 欄位名稱 | 說明 |
        |---|---|
        | 客戶 | 客戶代號或客戶編號 |
        | 客戶簡稱 | 客戶顯示名稱 |
        | 1月、2月、3月... | 每月銷售金額 |

        客戶資料統計為選填。若使用者自行上傳客戶資料統計檔，程式會嘗試帶入：聯絡人、聯絡電話、聯絡人手機；若未上傳，報表中的聯絡資料欄位將保留空白。
        """
    )

# =====================================
# 分析設定區
# =====================================

st.markdown('<div class="section-title">⚙️ 分析設定</div>', unsafe_allow_html=True)

config_col1, config_col2 = st.columns(2)

with config_col1:
    start_month = st.selectbox(
        "起始月份",
        options=list(range(1, 13)),
        index=0,
        format_func=lambda x: f"{x}月",
        help="請選擇分析期間的起始月份。"
    )

with config_col2:
    end_month = st.selectbox(
        "結束月份",
        options=list(range(1, 13)),
        index=2,
        format_func=lambda x: f"{x}月",
        help="請選擇分析期間的結束月份。"
    )

if start_month > end_month:
    analysis_months = []
    st.error("起始月份不能大於結束月份，請重新選擇。")
else:
    analysis_months = list(range(start_month, end_month + 1))
    st.success(f"本次分析月份：{start_month}月 ～ {end_month}月，共 {len(analysis_months)} 個月。")

# 企業使用者不需要看到 VIP 斷層比例與平滑底數，因此這裡只在程式內固定使用。
# 若未來管理者想調整，可直接修改 DEFAULT_VIP_GAP_RATIO 與 DEFAULT_VIP_LOG_BASE。
setting_preview = {
    "分析月份": f"{start_month}月～{end_month}月" if analysis_months else "月份設定錯誤",
    "銷貨資料": sales_file.name if sales_file is not None else "尚未上傳",
    "客戶資料來源": customer_source_text,
    "VIP 判定與平滑參數": "系統預設，前台不顯示"
}

preview_col1, preview_col2 = st.columns([1.2, 1])
with preview_col1:
    st.markdown("#### 🔎 本次設定預覽")
    st.dataframe(
        pd.DataFrame(
            list(setting_preview.items()),
            columns=["項目", "內容"]
        ),
        use_container_width=True,
        hide_index=True
    )

with preview_col2:
    st.markdown("#### 🧾 報表內容")
    st.markdown(
        """
        - 群組摘要
        - RFM 評分規範
        - 圖表總覽
        - 客戶分群結果
        - VIP 客戶維護清單
        - 高價值客戶維護清單
        - 一般客戶經營清單
        - 流失風險客戶追蹤清單
        """
    )

# =====================================
# 產生報表區
# =====================================

st.markdown('<div class="section-title">🚀 產生報表</div>', unsafe_allow_html=True)

run_col1, run_col2 = st.columns([1, 2])

with run_col1:
    run_button = st.button("開始分析並產生報表", type="primary", use_container_width=True)

with run_col2:
    st.markdown(
        '<p class="small-note">按下後系統會開始讀取 Excel、計算 RFM、進行 K-means 分群並產生報表。資料量較大時請稍候。</p>',
        unsafe_allow_html=True
    )

if run_button:
    if sales_file is None:
        st.error("請先上傳銷貨資料 Excel。")

    elif len(analysis_months) == 0:
        st.error("請確認起始月份與結束月份設定正確。")

    else:
        progress_bar = st.progress(0)
        status_text = st.empty()

        try:
            status_text.info("正在讀取資料並進行客群分析...")
            progress_bar.progress(25)

            with st.spinner("報表產生中，請稍候..."):
                output_file = generate_customer_report(
                    sales_file=sales_file,
                    customer_info_file=customer_file,
                    analysis_months=analysis_months,
                    vip_gap_ratio_threshold=DEFAULT_VIP_GAP_RATIO,
                    vip_m_log_base=DEFAULT_VIP_LOG_BASE
                )

            progress_bar.progress(85)
            status_text.info("正在準備下載檔案...")

            with open(output_file, "rb") as f:
                report_bytes = f.read()

            progress_bar.progress(100)
            status_text.success("報表產生完成！")

            st.download_button(
                label="📥 下載客群分析結果報表",
                data=report_bytes,
                file_name=Path(output_file).name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            st.success("已完成分析。下載後請開啟 Excel 檢查各工作表內容。")

        except Exception as e:
            progress_bar.empty()
            status_text.empty()
            st.error("報表產生失敗，請檢查上傳檔案欄位格式是否正確。")
            st.exception(e)
